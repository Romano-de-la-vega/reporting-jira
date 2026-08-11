#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyse KPI d'un export RSS/XML Jira.

Usage :
    python analyse_jira_rss.py "ticket non clot (Jira)(1).xml"
    python analyse_jira_rss.py "jira.xml" --output "rapport_jira.xlsx" --as-of "2026-06-05 13:55:18+00:00"

Sorties :
    - un fichier Excel .xlsx si openpyxl est installé
    - sinon plusieurs CSV dans un dossier *_csv

Ce script ne modifie jamais le XML source.
"""

from __future__ import annotations

import argparse
import csv
import html
import os
import re
import sys
import statistics
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, timezone, date
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


# -----------------------------
# Nettoyage / parsing robuste
# -----------------------------

_EMPTY_VALUES = {"", "-", "n/a", "na", "none", "null", "non renseigné", "non renseigne", "not specified", "ns (not specified)"}

_TEMPLATE_WORDS = {
    "context", "as a", "i want", "so that", "acceptance criteria", "test 1", "given", "when", "then",
    "additional information", "cause", "cause description", "impact", "impact description",
    "if yes specify", "if yes, specify", "creation", "modification", "deletion", "other",
    "tad", "mocp", "role matrix", "ddic", "data model", "training documentation", "pivot", "talend flow",
}


@dataclass
class Ticket:
    key: str
    id: str
    link: str
    project_key: str
    project_name: str
    title: str
    summary: str
    issue_type: str
    priority: str
    status: str
    status_category: str
    resolution: str
    assignee: str
    reporter: str
    created: Optional[datetime]
    updated: Optional[datetime]
    resolved: Optional[datetime]
    due: Optional[datetime]
    labels: str
    parent: str
    description_text: str
    environment_text: str
    acceptance_criteria_text: str
    acceptance_criteria_source: str
    has_acceptance_criteria: bool
    has_description: bool
    is_unassigned: bool
    is_overdue: bool
    age_days: Optional[float]
    days_since_update: Optional[float]
    days_to_due: Optional[float]
    lead_time_days: Optional[float]
    minutes_creation_to_last_update: Optional[float]
    comments_count: int
    last_comment_date: Optional[datetime]
    last_comment_author: str
    days_since_last_comment: Optional[float]
    attachments_count: int
    attachments_total_size_mb: float
    first_attachment_date: Optional[datetime]
    last_attachment_date: Optional[datetime]
    issue_links_count: int
    inward_links_count: int
    outward_links_count: int
    subtasks_count: int
    customfields_filled_count: int
    customfields_total_count: int
    customfields_fill_rate: Optional[float]
    risk_level: str
    risk_priority: str
    sprint: str
    target_type: str
    taille_demande: str
    theme: str
    horizon: str
    sla_applicable: str
    sync_billable_part: str
    quality_score: int
    quality_flags: str
    likely_stale_30d: bool
    likely_stale_60d: bool
    likely_stale_90d: bool


def text_of(el: Optional[ET.Element]) -> str:
    """Texte concaténé d'un élément XML."""
    if el is None:
        return ""
    return "".join(el.itertext())


def unescape_repeat(value: str, max_rounds: int = 5) -> str:
    """Décodage HTML répété pour gérer &amp;#233; puis &#233;."""
    if value is None:
        return ""
    previous = str(value)
    for _ in range(max_rounds):
        current = html.unescape(previous)
        if current == previous:
            break
        previous = current
    return previous


def clean_rich_text(value: str) -> str:
    """Transforme le HTML/Jira encodé en texte lisible."""
    value = unescape_repeat(value or "")
    value = re.sub(r"(?i)<\s*br\s*/?\s*>", "\n", value)
    value = re.sub(r"(?i)</\s*(p|li|tr|th|td|div|h\d|ul|ol|table)\s*>", "\n", value)
    value = re.sub(r"<[^>]+>", " ", value)
    value = unescape_repeat(value)
    value = value.replace("\xa0", " ")
    value = re.sub(r"[ \t\r\f\v]+", " ", value)
    value = re.sub(r"\n\s*\n+", "\n", value)
    return value.strip()


def normalize_for_check(value: str) -> str:
    value = clean_rich_text(value).lower()
    value = re.sub(r"[’'`´]", "'", value)
    value = re.sub(r"[^a-z0-9àâäéèêëîïôöùûüçœæ' ]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def is_empty_value(value: str) -> bool:
    cleaned = normalize_for_check(value)
    return cleaned in _EMPTY_VALUES


def is_meaningful_text(value: str, min_chars: int = 25) -> bool:
    """Évite de considérer les templates Jira vides comme du contenu réel."""
    cleaned = normalize_for_check(value)
    if cleaned in _EMPTY_VALUES:
        return False

    reduced = cleaned
    for word in sorted(_TEMPLATE_WORDS, key=len, reverse=True):
        reduced = reduced.replace(word, " ")
    reduced = re.sub(r"\b(test|scenario|scénario|etant donné|étant donné|alors|quand)\b", " ", reduced)
    reduced = re.sub(r"\s+", " ", reduced).strip()

    # au moins quelques mots significatifs et pas seulement des intitulés
    words = [w for w in reduced.split() if len(w) >= 3]
    return len(reduced) >= min_chars and len(words) >= 4


def parse_date(value: str) -> Optional[datetime]:
    value = (value or "").strip()
    if not value:
        return None

    # Format Jira RSS : Fri, 5 Jun 2026 15:46:08 +0200
    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        pass

    # ISO ou formats fréquents
    for fmt in ("%Y-%m-%d %H:%M:%S%z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(value, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            continue
    return None


def parse_as_of(value: Optional[str], raw_xml: str) -> datetime:
    if value:
        parsed = parse_date(value)
        if parsed:
            return parsed
        # accepte aussi 2026-06-05 13:55:18+00:00 via fromisoformat
        try:
            dt = datetime.fromisoformat(value)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            raise SystemExit(f"Date --as-of non comprise: {value}")

    # RSS generated by JIRA ... at Fri Jun 05 13:55:18 UTC 2026
    m = re.search(r"generated by JIRA.*? at ([A-Z][a-z]{2} [A-Z][a-z]{2} \d{2} \d{2}:\d{2}:\d{2} UTC \d{4})", raw_xml, re.S)
    if m:
        try:
            return datetime.strptime(m.group(1), "%a %b %d %H:%M:%S UTC %Y").replace(tzinfo=timezone.utc)
        except Exception:
            pass

    return datetime.now(timezone.utc)


def days_between(start: Optional[datetime], end: Optional[datetime]) -> Optional[float]:
    if not start or not end:
        return None
    return round((end.astimezone(timezone.utc) - start.astimezone(timezone.utc)).total_seconds() / 86400, 2)


def minutes_between(start: Optional[datetime], end: Optional[datetime]) -> Optional[float]:
    if not start or not end:
        return None
    return round((end.astimezone(timezone.utc) - start.astimezone(timezone.utc)).total_seconds() / 60, 2)


def safe_float(value: str, default: float = 0.0) -> float:
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return default


def get_child_text(item: ET.Element, tag: str) -> str:
    return clean_rich_text(text_of(item.find(tag)))


def get_child_raw_text(item: ET.Element, tag: str) -> str:
    return unescape_repeat(text_of(item.find(tag))).strip()


def get_attr(el: Optional[ET.Element], attr: str) -> str:
    return "" if el is None else str(el.attrib.get(attr, ""))


def extract_project(item: ET.Element) -> Tuple[str, str]:
    el = item.find("project")
    return get_attr(el, "key"), clean_rich_text(text_of(el))


def extract_labels(item: ET.Element) -> str:
    labels_el = item.find("labels")
    if labels_el is None:
        return ""
    labels = [clean_rich_text(text_of(x)) for x in labels_el.findall("label")]
    labels = [x for x in labels if x]
    return "; ".join(labels)


def extract_customfields(item: ET.Element) -> Dict[str, Dict[str, Any]]:
    """
    Retourne dict par nom de champ :
    {
      'Acceptance Criteria': {'id': 'customfield_10056', 'key': '...', 'values': [...], 'text': '...'}
    }
    """
    result: Dict[str, Dict[str, Any]] = {}
    for cf in item.findall("./customfields/customfield"):
        name = clean_rich_text(text_of(cf.find("customfieldname"))).strip()
        if not name:
            name = cf.attrib.get("id", "unknown_customfield")

        values = []
        for val_el in cf.findall("./customfieldvalues/customfieldvalue"):
            txt = clean_rich_text(text_of(val_el))
            if txt:
                values.append(txt)

        result[name] = {
            "id": cf.attrib.get("id", ""),
            "key": cf.attrib.get("key", ""),
            "values": values,
            "text": "; ".join(values).strip(),
        }
    return result


def get_customfield(customfields: Dict[str, Dict[str, Any]], *names: str) -> str:
    names_lower = {n.lower() for n in names}
    for name, payload in customfields.items():
        if name.lower() in names_lower:
            return payload.get("text", "")
    return ""


def find_story_acceptance_criteria(story_text: str) -> str:
    """Extrait la section Acceptance criteria depuis Story Description si possible."""
    text = clean_rich_text(story_text)
    if not text:
        return ""
    lower = text.lower()
    patterns = ["acceptance criteria", "critères d'acceptation", "criteres d'acceptation", "critère d'acceptation", "critere d'acceptation"]
    start_idx = -1
    start_pat = ""
    for p in patterns:
        idx = lower.find(p)
        if idx != -1 and (start_idx == -1 or idx < start_idx):
            start_idx = idx
            start_pat = p
    if start_idx == -1:
        return ""

    start = start_idx + len(start_pat)
    tail = text[start:]
    end_markers = ["\nTest 1", "\nGiven", "\nWhen", "\nThen", "\nAdditional information", "\nInformations complémentaires"]
    end_positions = [tail.lower().find(m.lower()) for m in end_markers if tail.lower().find(m.lower()) != -1]
    if end_positions:
        tail = tail[: min(end_positions)]
    return tail.strip(" :\n\t")


def extract_acceptance_criteria(customfields: Dict[str, Dict[str, Any]], description_text: str) -> Tuple[str, str, bool]:
    # 1) champ dédié
    dedicated = get_customfield(customfields, "Acceptance Criteria", "Acceptance criteria", "Critères d'acceptation", "Criteres d'acceptation")
    if is_meaningful_text(dedicated):
        return dedicated, "customfield: Acceptance Criteria", True

    # 2) story description, mais uniquement si section vraiment remplie
    story = get_customfield(customfields, "Story Description", "Description Story")
    story_ac = find_story_acceptance_criteria(story)
    if is_meaningful_text(story_ac):
        return story_ac, "customfield: Story Description / Acceptance criteria", True

    # 3) description générale : résultats attendus / expected result / Gherkin
    desc_norm = normalize_for_check(description_text)
    has_expected_marker = any(x in desc_norm for x in [
        "acceptance criteria", "critere d'acceptation", "criteres d'acceptation", "résultats attendus", "resultats attendus",
        "expected result", "given", "when", "then", "etant donne", "étant donné",
    ])
    if has_expected_marker and is_meaningful_text(description_text, min_chars=80):
        return description_text, "description: expected/acceptance markers", True

    if dedicated:
        return dedicated, "customfield: Acceptance Criteria empty/template", False
    if story:
        return story_ac or story, "customfield: Story Description empty/template", False
    return "", "missing", False


def extract_comments(item: ET.Element) -> List[Dict[str, Any]]:
    comments = []
    for comment in item.findall("./comments/comment"):
        comments.append({
            "id": comment.attrib.get("id", ""),
            "author": comment.attrib.get("author", ""),
            "created": parse_date(comment.attrib.get("created", "")),
            "body": clean_rich_text(text_of(comment)),
        })
    return comments


def extract_attachments(item: ET.Element) -> List[Dict[str, Any]]:
    atts = []
    for att in item.findall("./attachments/attachment"):
        atts.append({
            "id": att.attrib.get("id", ""),
            "name": att.attrib.get("name", ""),
            "size": safe_float(att.attrib.get("size", "0"), 0.0),
            "author": att.attrib.get("author", ""),
            "created": parse_date(att.attrib.get("created", "")),
        })
    return atts


def extract_issue_links(item: ET.Element) -> Tuple[int, int, int]:
    inward = 0
    outward = 0
    for linktype in item.findall("./issuelinks/issuelinktype"):
        inward += len(linktype.findall("./inwardlinks/issuelink"))
        outward += len(linktype.findall("./outwardlinks/issuelink"))
    return inward + outward, inward, outward


def extract_subtasks_count(item: ET.Element) -> int:
    return len(item.findall("./subtasks/subtask"))


def quality_flags(ticket_dict: Dict[str, Any]) -> Tuple[int, List[str]]:
    flags = []
    score = 100

    def penalize(condition: bool, points: int, label: str) -> None:
        nonlocal score
        if condition:
            score -= points
            flags.append(label)

    penalize(not ticket_dict["has_acceptance_criteria"], 25, "AC manquants ou template")
    penalize(not ticket_dict["has_description"], 15, "Description vide")
    penalize(ticket_dict["is_unassigned"], 15, "Non assigné")
    penalize(ticket_dict["is_overdue"], 15, "Échéance dépassée")
    penalize((ticket_dict["days_since_update"] is not None and ticket_dict["days_since_update"] >= 30), 10, "Aucune mise à jour >=30j")
    penalize((ticket_dict["comments_count"] == 0), 5, "Aucun commentaire")
    penalize((ticket_dict["attachments_count"] == 0 and ticket_dict["issue_type"].lower() == "bug"), 5, "Bug sans pièce jointe")
    penalize((ticket_dict["priority"].lower() in {"highest", "high"} and ticket_dict["days_since_update"] is not None and ticket_dict["days_since_update"] >= 7), 10, "Priorité haute inactive >=7j")

    return max(score, 0), flags


# -----------------------------
# Analyse principale
# -----------------------------

def parse_tickets(xml_path: Path, as_of: datetime) -> Tuple[List[Ticket], List[Dict[str, Any]], List[Dict[str, Any]]]:
    parser = ET.XMLParser(encoding="utf-8")
    tree = ET.parse(xml_path, parser=parser)
    root = tree.getroot()

    tickets: List[Ticket] = []
    comments_flat: List[Dict[str, Any]] = []
    customfields_flat: List[Dict[str, Any]] = []

    for item in root.findall("./channel/item"):
        key_el = item.find("key")
        key = clean_rich_text(text_of(key_el))
        item_id = get_attr(key_el, "id")
        project_key, project_name = extract_project(item)

        customfields = extract_customfields(item)
        for cf_name, payload in customfields.items():
            customfields_flat.append({
                "ticket_key": key,
                "customfield_name": cf_name,
                "customfield_id": payload.get("id", ""),
                "customfield_key": payload.get("key", ""),
                "value": payload.get("text", ""),
                "is_filled": is_meaningful_text(payload.get("text", ""), min_chars=2),
            })

        description_text = clean_rich_text(text_of(item.find("description")))
        environment_text = clean_rich_text(text_of(item.find("environment")))
        ac_text, ac_source, has_ac = extract_acceptance_criteria(customfields, description_text)

        created = parse_date(text_of(item.find("created")))
        updated = parse_date(text_of(item.find("updated")))
        resolved = parse_date(text_of(item.find("resolved")))
        due = parse_date(text_of(item.find("due")))

        comments = extract_comments(item)
        for c in comments:
            comments_flat.append({
                "ticket_key": key,
                "comment_id": c["id"],
                "comment_author": c["author"],
                "comment_created": c["created"],
                "comment_body": c["body"],
            })

        comment_dates = [c["created"] for c in comments if c.get("created")]
        last_comment_date = max(comment_dates) if comment_dates else None
        last_comment_author = ""
        if last_comment_date:
            last = [c for c in comments if c.get("created") == last_comment_date]
            last_comment_author = last[-1].get("author", "") if last else ""

        attachments = extract_attachments(item)
        att_dates = [a["created"] for a in attachments if a.get("created")]
        issue_links_count, inward_links_count, outward_links_count = extract_issue_links(item)

        cf_total = len(customfields)
        cf_filled = sum(1 for p in customfields.values() if is_meaningful_text(p.get("text", ""), min_chars=2))
        cf_rate = round(cf_filled / cf_total, 4) if cf_total else None

        assignee = get_child_text(item, "assignee")
        priority = get_child_text(item, "priority")
        status = get_child_text(item, "status")
        status_cat_el = item.find("statusCategory")
        status_category = status_cat_el.attrib.get("key", "") if status_cat_el is not None else ""
        resolution = get_child_text(item, "resolution")
        issue_type = get_child_text(item, "type")

        ticket_base = {
            "key": key,
            "issue_type": issue_type,
            "priority": priority,
            "status": status,
            "has_acceptance_criteria": has_ac,
            "has_description": is_meaningful_text(description_text, min_chars=20),
            "is_unassigned": normalize_for_check(assignee) in {"non assigné", "non assigne", "unassigned"} or get_attr(item.find("assignee"), "accountid") == "-1",
            "is_overdue": bool(due and due.date() < as_of.date() and status_category != "done"),
            "days_since_update": days_between(updated, as_of),
            "comments_count": len(comments),
            "attachments_count": len(attachments),
        }
        score, flags = quality_flags(ticket_base)

        ticket = Ticket(
            key=key,
            id=item_id,
            link=get_child_raw_text(item, "link"),
            project_key=project_key,
            project_name=project_name,
            title=get_child_text(item, "title"),
            summary=get_child_text(item, "summary"),
            issue_type=issue_type,
            priority=priority,
            status=status,
            status_category=status_category,
            resolution=resolution,
            assignee=assignee,
            reporter=get_child_text(item, "reporter"),
            created=created,
            updated=updated,
            resolved=resolved,
            due=due,
            labels=extract_labels(item),
            parent=get_child_text(item, "parent"),
            description_text=description_text,
            environment_text=environment_text,
            acceptance_criteria_text=clean_rich_text(ac_text),
            acceptance_criteria_source=ac_source,
            has_acceptance_criteria=has_ac,
            has_description=is_meaningful_text(description_text, min_chars=20),
            is_unassigned=ticket_base["is_unassigned"],
            is_overdue=ticket_base["is_overdue"],
            age_days=days_between(created, as_of),
            days_since_update=days_between(updated, as_of),
            days_to_due=days_between(as_of, due) if due else None,
            lead_time_days=days_between(created, resolved) if resolved else None,
            minutes_creation_to_last_update=minutes_between(created, updated),
            comments_count=len(comments),
            last_comment_date=last_comment_date,
            last_comment_author=last_comment_author,
            days_since_last_comment=days_between(last_comment_date, as_of) if last_comment_date else None,
            attachments_count=len(attachments),
            attachments_total_size_mb=round(sum(a["size"] for a in attachments) / (1024 * 1024), 3),
            first_attachment_date=min(att_dates) if att_dates else None,
            last_attachment_date=max(att_dates) if att_dates else None,
            issue_links_count=issue_links_count,
            inward_links_count=inward_links_count,
            outward_links_count=outward_links_count,
            subtasks_count=extract_subtasks_count(item),
            customfields_filled_count=cf_filled,
            customfields_total_count=cf_total,
            customfields_fill_rate=cf_rate,
            risk_level=get_customfield(customfields, "Risk level"),
            risk_priority=get_customfield(customfields, "Risk priority"),
            sprint=get_customfield(customfields, "Sprint"),
            target_type=get_customfield(customfields, "Target Type"),
            taille_demande=get_customfield(customfields, "Taille de la demande"),
            theme=get_customfield(customfields, "Theme"),
            horizon=get_customfield(customfields, "Horizon"),
            sla_applicable=get_customfield(customfields, "SLA applicable"),
            sync_billable_part=get_customfield(customfields, "Sync billable part"),
            quality_score=score,
            quality_flags="; ".join(flags),
            likely_stale_30d=bool(days_between(updated, as_of) is not None and days_between(updated, as_of) >= 30),
            likely_stale_60d=bool(days_between(updated, as_of) is not None and days_between(updated, as_of) >= 60),
            likely_stale_90d=bool(days_between(updated, as_of) is not None and days_between(updated, as_of) >= 90),
        )
        tickets.append(ticket)

    return tickets, comments_flat, customfields_flat


def dt_to_excel(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def ticket_to_row(ticket: Ticket) -> Dict[str, Any]:
    row = asdict(ticket)
    for k, v in list(row.items()):
        row[k] = dt_to_excel(v)
    return row


def pct(n: int, d: int) -> float:
    return round(n / d, 4) if d else 0.0


def avg(values: Iterable[Optional[float]]) -> Optional[float]:
    vals = [float(v) for v in values if v is not None]
    return round(sum(vals) / len(vals), 2) if vals else None


def median(values: Iterable[Optional[float]]) -> Optional[float]:
    vals = sorted(float(v) for v in values if v is not None)
    return round(statistics.median(vals), 2) if vals else None


def p90(values: Iterable[Optional[float]]) -> Optional[float]:
    vals = sorted(float(v) for v in values if v is not None)
    if not vals:
        return None
    idx = int(round(0.9 * (len(vals) - 1)))
    return round(vals[idx], 2)


def compute_dashboard(tickets: List[Ticket], as_of: datetime) -> List[Dict[str, Any]]:
    total = len(tickets)
    open_tickets = [t for t in tickets if t.status_category != "done"]
    resolved = [t for t in tickets if t.resolved is not None or t.status_category == "done"]
    bugs = [t for t in tickets if t.issue_type.lower() == "bug"]
    stories = [t for t in tickets if t.issue_type.lower() == "story"]
    high_priority = [t for t in tickets if t.priority.lower() in {"highest", "high"}]

    rows = [
        {"KPI": "Date d'analyse", "Valeur": as_of.isoformat(), "Commentaire": "Date utilisée pour calculer les âges et retards"},
        {"KPI": "Tickets exportés", "Valeur": total, "Commentaire": "Nombre de balises <item>"},
        {"KPI": "Tickets non terminés", "Valeur": len(open_tickets), "Commentaire": "statusCategory différent de done"},
        {"KPI": "Tickets terminés / résolus", "Valeur": len(resolved), "Commentaire": "statusCategory done ou date resolved présente"},
        {"KPI": "Bugs", "Valeur": len(bugs), "Commentaire": "type = Bug"},
        {"KPI": "Stories", "Valeur": len(stories), "Commentaire": "type = Story"},
        {"KPI": "Tickets sans Acceptance Criteria", "Valeur": sum(not t.has_acceptance_criteria for t in tickets), "Commentaire": "Champ AC absent, vide ou template"},
        {"KPI": "% sans Acceptance Criteria", "Valeur": pct(sum(not t.has_acceptance_criteria for t in tickets), total), "Commentaire": "À surveiller en cadrage"},
        {"KPI": "Tickets sans description exploitable", "Valeur": sum(not t.has_description for t in tickets), "Commentaire": "Description vide/template"},
        {"KPI": "Tickets non assignés", "Valeur": sum(t.is_unassigned for t in tickets), "Commentaire": "Assignee = Non assigné / Unassigned"},
        {"KPI": "Tickets en retard", "Valeur": sum(t.is_overdue for t in tickets), "Commentaire": "Due date dépassée et non done"},
        {"KPI": "Tickets sans commentaire", "Valeur": sum(t.comments_count == 0 for t in tickets), "Commentaire": "Peut indiquer peu de traçabilité"},
        {"KPI": "Tickets sans pièce jointe", "Valeur": sum(t.attachments_count == 0 for t in tickets), "Commentaire": "À nuancer selon le type"},
        {"KPI": "Bugs sans pièce jointe", "Valeur": sum(t.attachments_count == 0 for t in bugs), "Commentaire": "Souvent problématique pour reproduction"},
        {"KPI": "Tickets sans lien Jira", "Valeur": sum(t.issue_links_count == 0 for t in tickets), "Commentaire": "Traçabilité fonctionnelle faible"},
        {"KPI": "Priorités High/Highest", "Valeur": len(high_priority), "Commentaire": "Charge critique"},
        {"KPI": "Âge moyen des tickets ouverts (jours)", "Valeur": avg(t.age_days for t in open_tickets), "Commentaire": "Depuis created jusqu'à date d'analyse"},
        {"KPI": "Âge médian des tickets ouverts (jours)", "Valeur": median(t.age_days for t in open_tickets), "Commentaire": "Plus robuste aux vieux tickets"},
        {"KPI": "P90 âge tickets ouverts (jours)", "Valeur": p90(t.age_days for t in open_tickets), "Commentaire": "90% des tickets ouverts sont sous ce seuil"},
        {"KPI": "Jours moyens depuis dernière mise à jour", "Valeur": avg(t.days_since_update for t in tickets), "Commentaire": "Stagnation potentielle"},
        {"KPI": "Tickets sans mise à jour >=30j", "Valeur": sum(t.likely_stale_30d for t in tickets), "Commentaire": "Risque d'obsolescence"},
        {"KPI": "Tickets sans mise à jour >=60j", "Valeur": sum(t.likely_stale_60d for t in tickets), "Commentaire": "Risque fort"},
        {"KPI": "Tickets sans mise à jour >=90j", "Valeur": sum(t.likely_stale_90d for t in tickets), "Commentaire": "Risque très fort"},
        {"KPI": "Score qualité moyen", "Valeur": avg(t.quality_score for t in tickets), "Commentaire": "100 = complet ; pénalités AC/desc/assignation/retard/inactivité"},
    ]
    return rows


def group_summary(tickets: List[Ticket], field: str) -> List[Dict[str, Any]]:
    groups: Dict[str, List[Ticket]] = defaultdict(list)
    for t in tickets:
        value = str(getattr(t, field) or "(vide)")
        groups[value].append(t)

    rows = []
    for value, items in sorted(groups.items(), key=lambda kv: (-len(kv[1]), kv[0].lower())):
        rows.append({
            field: value,
            "tickets": len(items),
            "sans_acceptance_criteria": sum(not t.has_acceptance_criteria for t in items),
            "pct_sans_acceptance_criteria": pct(sum(not t.has_acceptance_criteria for t in items), len(items)),
            "non_assignes": sum(t.is_unassigned for t in items),
            "en_retard": sum(t.is_overdue for t in items),
            "sans_maj_30j": sum(t.likely_stale_30d for t in items),
            "age_moyen_jours": avg(t.age_days for t in items),
            "age_median_jours": median(t.age_days for t in items),
            "jours_moyens_depuis_update": avg(t.days_since_update for t in items),
            "score_qualite_moyen": avg(t.quality_score for t in items),
        })
    return rows


def quality_alerts(tickets: List[Ticket]) -> List[Dict[str, Any]]:
    rows = []
    for t in sorted(tickets, key=lambda x: (x.quality_score, -(x.age_days or 0))):
        if t.quality_flags:
            rows.append({
                "key": t.key,
                "summary": t.summary,
                "link": t.link,
                "type": t.issue_type,
                "priority": t.priority,
                "status": t.status,
                "assignee": t.assignee,
                "created": dt_to_excel(t.created),
                "updated": dt_to_excel(t.updated),
                "age_days": t.age_days,
                "days_since_update": t.days_since_update,
                "quality_score": t.quality_score,
                "quality_flags": t.quality_flags,
            })
    return rows


def customfield_coverage(customfields_flat: List[Dict[str, Any]], tickets_count: int) -> List[Dict[str, Any]]:
    by_field: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in customfields_flat:
        by_field[row["customfield_name"]].append(row)

    rows = []
    for name, rows_cf in sorted(by_field.items(), key=lambda kv: kv[0].lower()):
        filled = sum(bool(r["is_filled"]) for r in rows_cf)
        rows.append({
            "customfield_name": name,
            "customfield_id": rows_cf[0].get("customfield_id", ""),
            "customfield_key": rows_cf[0].get("customfield_key", ""),
            "present_on_tickets": len(rows_cf),
            "filled_count": filled,
            "empty_count": len(rows_cf) - filled,
            "fill_rate_vs_present": pct(filled, len(rows_cf)),
            "fill_rate_vs_all_tickets": pct(filled, tickets_count),
        })
    return rows


def write_csvs(output_base: Path, sheets: Dict[str, List[Dict[str, Any]]]) -> None:
    out_dir = output_base.with_suffix("").as_posix() + "_csv"
    os.makedirs(out_dir, exist_ok=True)
    for name, rows in sheets.items():
        path = Path(out_dir) / f"{name}.csv"
        fieldnames = sorted({k for row in rows for k in row.keys()}) if rows else ["empty"]
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                cleaned = {k: dt_to_excel(v) for k, v in row.items()}
                writer.writerow(cleaned)
    print(f"openpyxl indisponible : CSV écrits dans {out_dir}")


def write_excel(output_path: Path, sheets: Dict[str, List[Dict[str, Any]]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.chart import BarChart, PieChart, Reference
    except Exception:
        write_csvs(output_path, sheets)
        return

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            ws.append(["Aucune donnée"])
            continue

        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in rows:
            ws.append([dt_to_excel(row.get(h)) for h in headers])

        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Formats simples
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            header_lower = header.lower()
            max_len = len(str(header))
            for cell in ws[col_letter][1: min(ws.max_row, 1000)]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value))[:0] if False else len(str(cell.value)))
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
            width = min(max(max_len + 2, 10), 55)
            if any(x in header_lower for x in ["description", "summary", "title", "criteria", "flags", "comment", "link"]):
                width = min(max(width, 25), 70)
            ws.column_dimensions[col_letter].width = width

        # Table Excel
        if ws.max_row >= 2 and ws.max_column >= 2:
            table_name = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name)[:20] + "Tbl"
            table_ref = ws.dimensions
            tab = Table(displayName=table_name, ref=table_ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            try:
                ws.add_table(tab)
            except Exception:
                pass

    # Graphiques simples dans Dashboard
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 70

    # Ajoute quelques graphiques sur sheets groupées
    for sheet_name in ["Par_status", "Par_assignee", "Par_type", "Par_priority"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row >= 3:
                chart = BarChart()
                chart.title = f"Tickets - {sheet_name}"
                chart.y_axis.title = "Nombre de tickets"
                chart.x_axis.title = sheet_name.replace("Par_", "")
                data = Reference(ws, min_col=2, min_row=1, max_row=min(ws.max_row, 15))
                cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 15))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 7
                chart.width = 14
                ws.add_chart(chart, "M2")

    wb.save(output_path)
    print(f"Rapport écrit : {output_path}")


def main() -> None:
    argp = argparse.ArgumentParser(description="Analyse KPI d'un export RSS/XML Jira")
    argp.add_argument("xml", help="Chemin du fichier XML RSS Jira")
    argp.add_argument("--output", "-o", default="jira_kpi_report.xlsx", help="Fichier de sortie .xlsx")
    argp.add_argument("--as-of", default=None, help="Date d'analyse, ex: '2026-06-05 13:55:18+00:00'. Par défaut: date de génération RSS si trouvée, sinon maintenant.")
    args = argp.parse_args()

    xml_path = Path(args.xml)
    if not xml_path.exists():
        raise SystemExit(f"Fichier introuvable: {xml_path}")

    raw_xml = xml_path.read_text(encoding="utf-8", errors="replace")
    as_of = parse_as_of(args.as_of, raw_xml)

    tickets, comments_flat, customfields_flat = parse_tickets(xml_path, as_of)
    ticket_rows = [ticket_to_row(t) for t in tickets]

    # Conversion datetime pour les tables plates
    for row in comments_flat:
        row["comment_created"] = dt_to_excel(row.get("comment_created"))

    sheets = {
        "Dashboard": compute_dashboard(tickets, as_of),
        "Tickets": ticket_rows,
        "Alertes_qualite": quality_alerts(tickets),
        "Par_status": group_summary(tickets, "status"),
        "Par_assignee": group_summary(tickets, "assignee"),
        "Par_type": group_summary(tickets, "issue_type"),
        "Par_priority": group_summary(tickets, "priority"),
        "Par_sprint": group_summary(tickets, "sprint"),
        "Par_theme": group_summary(tickets, "theme"),
        "CustomFields": customfields_flat,
        "Couverture_champs": customfield_coverage(customfields_flat, len(tickets)),
        "Commentaires": comments_flat,
    }

    output_path = Path(args.output)
    write_excel(output_path, sheets)

    # Résumé console
    total = len(tickets)
    sans_ac = sum(not t.has_acceptance_criteria for t in tickets)
    non_assignes = sum(t.is_unassigned for t in tickets)
    stale_30 = sum(t.likely_stale_30d for t in tickets)
    print("\nRésumé rapide")
    print("--------------")
    print(f"Tickets analysés              : {total}")
    print(f"Sans Acceptance Criteria      : {sans_ac} ({pct(sans_ac, total):.1%})")
    print(f"Non assignés                  : {non_assignes} ({pct(non_assignes, total):.1%})")
    print(f"Sans mise à jour >= 30 jours  : {stale_30} ({pct(stale_30, total):.1%})")
    print("\nLimite importante : le RSS Jira ne contient généralement pas le changelog complet champ par champ.")
    print("Le script trace donc les signaux disponibles dans ce fichier : updated, commentaires, pièces jointes et liens.")


if __name__ == "__main__":
    main()
