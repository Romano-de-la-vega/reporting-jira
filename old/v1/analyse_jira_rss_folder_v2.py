#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyse des modifications entre plusieurs exports RSS/XML Jira.

Objectif :
    Donner des KPI fiables de modification entre plusieurs snapshots Jira exportes en RSS/XML :
    descriptions modifiees, acceptance criteria modifies, changements de statut,
    changements d'assigne, apparitions/disparitions de tickets, etc.

Principe important :
    Le RSS Jira ne contient pas le changelog complet champ par champ.
    Le script compare donc les snapshots entre eux. Une modification detectee entre
    deux fichiers signifie : "le ticket est different entre l'export precedent et
    l'export courant". La date exacte et l'auteur exact de la modification ne sont
    pas disponibles dans ce format RSS.

Comparaison des tickets :
    Le script compare les tickets sur l'id technique Jira : <key id="...">.
    Si cet id est absent, il utilise la key Jira comme fallback et le signale.

Usage :
    python analyse_jira_rss.py "C:/chemin/vers/repertoire_exports" --output jira_kpi_report.xlsx

    python analyse_jira_rss.py "C:/chemin/vers/repertoire_exports" \
        --period-start "2026-06-01" \
        --period-end "2026-06-05" \
        --sprint "Sprint 16" \
        --status "Ouvert" \
        --output jira_kpi_report.xlsx

Dependance :
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import hashlib
import html
import os
import re
import statistics
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, time, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


# =============================================================================
# Configuration
# =============================================================================

EMPTY_VALUES = {
    "", "-", "n/a", "na", "none", "null", "non renseigne", "non renseigné",
    "not specified", "ns not specified", "ns (not specified)", "aucun", "vide",
}

TEMPLATE_WORDS = {
    "context", "as a", "i want", "so that", "acceptance criteria", "test 1",
    "given", "when", "then", "additional information", "cause", "cause description",
    "impact", "impact description", "if yes specify", "if yes, specify", "creation",
    "modification", "deletion", "other", "tad", "mocp", "role matrix", "ddic",
    "data model", "training documentation", "pivot", "talend flow",
}

# Champs simples compares. Les gros textes sont traites a part.
TRACKED_SIMPLE_FIELDS = [
    ("summary", "Résumé"),
    ("issue_type", "Type"),
    ("priority", "Priorité"),
    ("status", "Statut"),
    ("status_category", "Catégorie statut"),
    ("assignee", "Assigné"),
    ("reporter", "Reporter"),
    ("sprint", "Sprint"),
    ("theme", "Thème"),
    ("risk_level", "Risk level"),
    ("risk_priority", "Risk priority"),
    ("target_type", "Target Type"),
    ("taille_demande", "Taille de la demande"),
    ("horizon", "Horizon"),
    ("sla_applicable", "SLA applicable"),
    ("due_iso", "Due date"),
    ("labels", "Labels"),
    ("resolution", "Résolution"),
]

# Champs custom a inclure explicitement si presents.
IMPORTANT_CUSTOMFIELDS = {
    "Acceptance Criteria",
    "Acceptance criteria",
    "Critères d'acceptation",
    "Criteres d'acceptation",
    "Story Description",
    "Sprint",
    "Theme",
    "Risk level",
    "Risk priority",
    "Target Type",
    "Taille de la demande",
    "Horizon",
    "SLA applicable",
    "Sync billable part",
}


# =============================================================================
# Modeles de donnees
# =============================================================================

@dataclass
class TicketSnapshot:
    ticket_id: str
    key: str
    key_id_raw: str
    file_name: str
    export_date: datetime
    link: str = ""
    project_key: str = ""
    project_name: str = ""
    summary: str = ""
    title: str = ""
    issue_type: str = ""
    priority: str = ""
    status: str = ""
    status_category: str = ""
    resolution: str = ""
    assignee: str = ""
    reporter: str = ""
    created: Optional[datetime] = None
    updated: Optional[datetime] = None
    due: Optional[datetime] = None
    labels: str = ""
    sprint: str = ""
    theme: str = ""
    risk_level: str = ""
    risk_priority: str = ""
    target_type: str = ""
    taille_demande: str = ""
    horizon: str = ""
    sla_applicable: str = ""
    sync_billable_part: str = ""
    description_text: str = ""
    acceptance_criteria_text: str = ""
    acceptance_criteria_source: str = ""
    has_acceptance_criteria: bool = False
    comments_count: int = 0
    comments_signature: str = ""
    attachments_count: int = 0
    attachments_signature: str = ""
    issue_links_count: int = 0
    issue_links_signature: str = ""
    subtasks_count: int = 0
    customfields_signature: str = ""
    customfields: Dict[str, str] = field(default_factory=dict)

    @property
    def due_iso(self) -> str:
        return format_dt(self.due, date_only=True)

    @property
    def description_hash(self) -> str:
        return stable_hash(canonical_text(self.description_text))

    @property
    def acceptance_criteria_hash(self) -> str:
        return stable_hash(canonical_text(self.acceptance_criteria_text))


@dataclass
class SnapshotFile:
    path: Path
    export_date: datetime
    source_date_text: str
    tickets: Dict[str, TicketSnapshot]
    warnings: List[str]


@dataclass
class ChangeEvent:
    window_start: datetime
    window_end: datetime
    previous_file: str
    current_file: str
    ticket_id: str
    key_before: str
    key_after: str
    project_key: str
    sprint_before: str
    sprint_after: str
    sprint_effective: str
    status_before: str
    status_after: str
    status_effective: str
    status_category_before: str
    status_category_after: str
    issue_type_before: str
    issue_type_after: str
    issue_type_effective: str
    assignee_before: str
    assignee_after: str
    priority_before: str
    priority_after: str
    field_name: str
    change_type: str
    before_value: str
    after_value: str
    before_length: Optional[int]
    after_length: Optional[int]
    before_excerpt: str
    after_excerpt: str
    jira_updated_before: Optional[datetime]
    jira_updated_after: Optional[datetime]
    link: str

    def as_row(self) -> Dict[str, Any]:
        return {
            "Fenêtre début": excel_dt(self.window_start),
            "Fenêtre fin / détection": excel_dt(self.window_end),
            "Fichier précédent": self.previous_file,
            "Fichier courant": self.current_file,
            "Ticket ID": self.ticket_id,
            "Key avant": self.key_before,
            "Key après": self.key_after,
            "Projet": self.project_key,
            "Sprint avant": self.sprint_before,
            "Sprint après": self.sprint_after,
            "Sprint effectif": self.sprint_effective,
            "Statut avant": self.status_before,
            "Statut après": self.status_after,
            "Statut effectif": self.status_effective,
            "Catégorie statut avant": self.status_category_before,
            "Catégorie statut après": self.status_category_after,
            "Type avant": self.issue_type_before,
            "Type après": self.issue_type_after,
            "Type effectif": self.issue_type_effective,
            "Assigné avant": self.assignee_before,
            "Assigné après": self.assignee_after,
            "Priorité avant": self.priority_before,
            "Priorité après": self.priority_after,
            "Champ modifié": self.field_name,
            "Type de modification": self.change_type,
            "Valeur avant": self.before_value,
            "Valeur après": self.after_value,
            "Taille avant": self.before_length,
            "Taille après": self.after_length,
            "Extrait avant": self.before_excerpt,
            "Extrait après": self.after_excerpt,
            "Jira updated avant": excel_dt(self.jira_updated_before),
            "Jira updated après": excel_dt(self.jira_updated_after),
            "Lien": self.link,
        }


# =============================================================================
# Nettoyage, dates, hashing
# =============================================================================

def unescape_repeat(value: str, max_rounds: int = 5) -> str:
    if value is None:
        return ""
    previous = str(value)
    for _ in range(max_rounds):
        current = html.unescape(previous)
        if current == previous:
            break
        previous = current
    return previous


def clean_rich_text(value: str) -> str:
    """Decode le HTML Jira encode dans le RSS et produit un texte lisible."""
    value = unescape_repeat(value or "")
    value = re.sub(r"(?i)<\s*br\s*/?\s*>", "\n", value)
    value = re.sub(r"(?i)</\s*(p|li|tr|th|td|div|h\d|ul|ol|table)\s*>", "\n", value)
    value = re.sub(r"<[^>]+>", " ", value)
    value = unescape_repeat(value)
    value = value.replace("\xa0", " ")
    value = re.sub(r"[ \t\r\f\v]+", " ", value)
    value = re.sub(r"\n\s*\n+", "\n", value)
    return value.strip()


def canonical_text(value: str) -> str:
    """Version stable pour comparaison : insensitive aux espaces parasites."""
    value = clean_rich_text(value)
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n\s+", "\n", value)
    value = re.sub(r"\s+\n", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def normalize_for_check(value: str) -> str:
    value = canonical_text(value).lower()
    value = re.sub(r"[’'`´]", "'", value)
    value = re.sub(r"[^a-z0-9àâäéèêëîïôöùûüçœæ' ]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def is_meaningful_text(value: str, min_chars: int = 25) -> bool:
    cleaned = normalize_for_check(value)
    if cleaned in EMPTY_VALUES:
        return False

    reduced = cleaned
    for word in sorted(TEMPLATE_WORDS, key=len, reverse=True):
        reduced = reduced.replace(word, " ")
    reduced = re.sub(r"\b(test|scenario|scénario|etant donné|étant donné|alors|quand)\b", " ", reduced)
    reduced = re.sub(r"\s+", " ", reduced).strip()
    words = [w for w in reduced.split() if len(w) >= 3]
    return len(reduced) >= min_chars and len(words) >= 4


def stable_hash(value: str) -> str:
    return hashlib.sha256((value or "").encode("utf-8", errors="replace")).hexdigest()[:16]


def excerpt(value: str, max_chars: int = 650) -> str:
    value = canonical_text(value)
    value = value.replace("\n", " | ")
    if len(value) <= max_chars:
        return value
    return value[: max_chars - 3].rstrip() + "..."


def text_of(el: Optional[ET.Element]) -> str:
    if el is None:
        return ""
    return "".join(el.itertext())


def attr(el: Optional[ET.Element], name: str) -> str:
    if el is None:
        return ""
    return str(el.attrib.get(name, ""))


def parse_jira_date(value: str) -> Optional[datetime]:
    value = (value or "").strip()
    if not value:
        return None

    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        pass

    # Formats frequents.
    for fmt in (
        "%Y-%m-%d %H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            dt = datetime.strptime(value, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except Exception:
            continue

    try:
        dt = datetime.fromisoformat(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def parse_period_boundary(value: Optional[str], end_of_day: bool = False) -> Optional[datetime]:
    if not value:
        return None
    raw = value.strip()
    parsed = parse_jira_date(raw)
    if parsed:
        # Si l'utilisateur donne seulement une date, parse_jira_date renvoie minuit.
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4}", raw):
            if end_of_day:
                return datetime.combine(parsed.date(), time.max, tzinfo=timezone.utc)
            return datetime.combine(parsed.date(), time.min, tzinfo=timezone.utc)
        return parsed
    raise SystemExit(f"Date non comprise: {value}")


def extract_export_date(raw_xml: str, file_path: Path) -> Tuple[datetime, str, List[str]]:
    warnings: List[str] = []

    # Exemple : RSS generated by JIRA (...) at Fri Jun 05 13:55:18 UTC 2026
    m = re.search(
        r"RSS generated by JIRA.*? at ([A-Z][a-z]{2} [A-Z][a-z]{2} \d{2} \d{2}:\d{2}:\d{2} UTC \d{4})",
        raw_xml,
        re.S,
    )
    if m:
        source = m.group(1).strip()
        try:
            dt = datetime.strptime(source, "%a %b %d %H:%M:%S UTC %Y").replace(tzinfo=timezone.utc)
            return dt, source, warnings
        except Exception:
            warnings.append(f"Date RSS generated by JIRA illisible dans {file_path.name}: {source}")

    # Fallback : date max updated/created dans le fichier.
    dates = []
    for tag in ("created", "updated", "resolved", "due"):
        for match in re.finditer(rf"<{tag}>(.*?)</{tag}>", raw_xml, flags=re.S):
            dt = parse_jira_date(clean_rich_text(match.group(1)))
            if dt:
                dates.append(dt)
    if dates:
        dt = max(dates)
        warnings.append(f"Date d'export absente dans {file_path.name}; fallback = date max Jira ({dt.isoformat()}).")
        return dt, "fallback:max_jira_date", warnings

    # Dernier fallback : date de modification du fichier local.
    dt = datetime.fromtimestamp(file_path.stat().st_mtime, tz=timezone.utc)
    warnings.append(f"Date d'export absente dans {file_path.name}; fallback = date fichier local ({dt.isoformat()}).")
    return dt, "fallback:file_mtime", warnings


def excel_dt(dt: Optional[datetime]) -> Optional[datetime]:
    if not dt:
        return None
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def format_dt(dt: Optional[datetime], date_only: bool = False) -> str:
    if not dt:
        return ""
    dt = dt.astimezone(timezone.utc)
    return dt.strftime("%Y-%m-%d") if date_only else dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def days_between(start: Optional[datetime], end: Optional[datetime]) -> Optional[float]:
    if not start or not end:
        return None
    return round((end.astimezone(timezone.utc) - start.astimezone(timezone.utc)).total_seconds() / 86400, 2)


def pct(n: int, d: int) -> float:
    return round(n / d, 4) if d else 0.0


def avg(values: Iterable[Optional[float]]) -> Optional[float]:
    vals = [float(v) for v in values if v is not None]
    return round(sum(vals) / len(vals), 2) if vals else None


def median(values: Iterable[Optional[float]]) -> Optional[float]:
    vals = sorted(float(v) for v in values if v is not None)
    return round(statistics.median(vals), 2) if vals else None


# =============================================================================
# Parsing XML Jira RSS
# =============================================================================

def find_xml_files(input_path: Path) -> List[Path]:
    if input_path.is_file():
        return [input_path]
    if not input_path.exists():
        raise SystemExit(f"Chemin introuvable: {input_path}")
    files = sorted([p for p in input_path.rglob("*.xml") if p.is_file()])
    if not files:
        raise SystemExit(f"Aucun fichier .xml trouve dans: {input_path}")
    return files


def get_child_text(item: ET.Element, tag: str) -> str:
    return clean_rich_text(text_of(item.find(tag)))


def get_child_raw(item: ET.Element, tag: str) -> str:
    return unescape_repeat(text_of(item.find(tag))).strip()


def extract_labels(item: ET.Element) -> str:
    labels_el = item.find("labels")
    if labels_el is None:
        return ""
    labels = [clean_rich_text(text_of(x)) for x in labels_el.findall("label")]
    return "; ".join([x for x in labels if x])


def extract_project(item: ET.Element) -> Tuple[str, str]:
    project = item.find("project")
    return attr(project, "key"), clean_rich_text(text_of(project))


def extract_customfields(item: ET.Element) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for cf in item.findall("./customfields/customfield"):
        name = clean_rich_text(text_of(cf.find("customfieldname"))).strip()
        if not name:
            name = attr(cf, "id") or "unknown_customfield"

        values = []
        for val_el in cf.findall("./customfieldvalues/customfieldvalue"):
            txt = clean_rich_text(text_of(val_el))
            if txt:
                values.append(txt)
        result[name] = "; ".join(values).strip()
    return result


def get_cf(customfields: Dict[str, str], *names: str) -> str:
    wanted = {x.lower() for x in names}
    for name, value in customfields.items():
        if name.lower() in wanted:
            return value
    return ""


def find_story_acceptance_criteria(story_text: str) -> str:
    text = canonical_text(story_text)
    if not text:
        return ""
    lower = text.lower()
    patterns = [
        "acceptance criteria",
        "critères d'acceptation",
        "criteres d'acceptation",
        "critère d'acceptation",
        "critere d'acceptation",
    ]
    start_idx = -1
    start_pat = ""
    for pattern in patterns:
        idx = lower.find(pattern)
        if idx != -1 and (start_idx == -1 or idx < start_idx):
            start_idx = idx
            start_pat = pattern
    if start_idx == -1:
        return ""

    tail = text[start_idx + len(start_pat):]
    tail_lower = tail.lower()
    end_markers = [
        "\ntest 1", "\ngiven", "\nwhen", "\nthen", "\nadditional information",
        "\ninformations complémentaires", "\ninformation complémentaire",
    ]
    positions = [tail_lower.find(m) for m in end_markers if tail_lower.find(m) != -1]
    if positions:
        tail = tail[: min(positions)]
    return tail.strip(" :\n\t")


def extract_acceptance_criteria(customfields: Dict[str, str], description_text: str) -> Tuple[str, str, bool]:
    dedicated = get_cf(
        customfields,
        "Acceptance Criteria",
        "Acceptance criteria",
        "Critères d'acceptation",
        "Criteres d'acceptation",
    )
    if dedicated:
        return canonical_text(dedicated), "customfield: Acceptance Criteria", is_meaningful_text(dedicated)

    story = get_cf(customfields, "Story Description", "Description Story")
    story_ac = find_story_acceptance_criteria(story)
    if story_ac:
        return canonical_text(story_ac), "customfield: Story Description / Acceptance criteria", is_meaningful_text(story_ac)

    # Fallback description : utile pour les tickets qui portent les criteres dans la description.
    desc_norm = normalize_for_check(description_text)
    markers = [
        "acceptance criteria", "critere d'acceptation", "criteres d'acceptation",
        "résultat attendu", "resultat attendu", "résultats attendus", "resultats attendus",
        "expected result", "given", "when", "then", "étant donné", "etant donne",
    ]
    if any(m in desc_norm for m in markers) and is_meaningful_text(description_text, min_chars=80):
        return canonical_text(description_text), "description: expected/acceptance markers", True

    return "", "missing", False


def extract_comments_signature(item: ET.Element) -> Tuple[int, str]:
    parts = []
    for c in item.findall("./comments/comment"):
        parts.append("|".join([
            attr(c, "id"),
            attr(c, "author"),
            attr(c, "created"),
            stable_hash(canonical_text(text_of(c))),
        ]))
    return len(parts), stable_hash("\n".join(sorted(parts)))


def extract_attachments_signature(item: ET.Element) -> Tuple[int, str]:
    parts = []
    for a in item.findall("./attachments/attachment"):
        parts.append("|".join([
            attr(a, "id"), attr(a, "name"), attr(a, "size"), attr(a, "created"), attr(a, "author")
        ]))
    return len(parts), stable_hash("\n".join(sorted(parts)))


def extract_issue_links_signature(item: ET.Element) -> Tuple[int, str]:
    parts = []
    for linktype in item.findall("./issuelinks/issuelinktype"):
        type_name = clean_rich_text(text_of(linktype.find("name")))
        for node in linktype.findall("./inwardlinks/issuelink/issuekey"):
            parts.append(f"in|{type_name}|{attr(node, 'id')}|{clean_rich_text(text_of(node))}")
        for node in linktype.findall("./outwardlinks/issuelink/issuekey"):
            parts.append(f"out|{type_name}|{attr(node, 'id')}|{clean_rich_text(text_of(node))}")
    return len(parts), stable_hash("\n".join(sorted(parts)))


def extract_subtasks_count(item: ET.Element) -> int:
    return len(item.findall("./subtasks/subtask"))


def important_customfields_signature(customfields: Dict[str, str]) -> str:
    parts = []
    for name, value in sorted(customfields.items(), key=lambda kv: kv[0].lower()):
        # On garde tous les champs avec une valeur, mais on normalise pour eviter les faux positifs d'espaces.
        if value or name in IMPORTANT_CUSTOMFIELDS:
            parts.append(f"{name}={canonical_text(value)}")
    return stable_hash("\n".join(parts))


def parse_snapshot_file(path: Path) -> SnapshotFile:
    raw = path.read_text(encoding="utf-8", errors="replace")
    export_date, source_date_text, warnings = extract_export_date(raw, path)

    try:
        root = ET.fromstring(raw.encode("utf-8", errors="replace"))
    except Exception as exc:
        raise SystemExit(f"XML illisible dans {path}: {exc}")

    tickets: Dict[str, TicketSnapshot] = {}
    seen_ids = Counter()

    for item in root.findall("./channel/item"):
        key_el = item.find("key")
        key = clean_rich_text(text_of(key_el))
        key_id_raw = attr(key_el, "id")
        ticket_id = key_id_raw or key
        if not ticket_id:
            warnings.append(f"Ticket sans key/id ignore dans {path.name}")
            continue
        if not key_id_raw:
            warnings.append(f"Fallback sur key car id technique absent: {key} dans {path.name}")

        seen_ids[ticket_id] += 1
        project_key = clean_rich_text(text_of(item.find("key")))
        project_name = extract_project(item)[1]
        customfields = extract_customfields(item)
        description_text = canonical_text(text_of(item.find("description")))
        ac_text, ac_source, has_ac = extract_acceptance_criteria(customfields, description_text)
        comments_count, comments_signature = extract_comments_signature(item)
        attachments_count, attachments_signature = extract_attachments_signature(item)
        links_count, links_signature = extract_issue_links_signature(item)
        status_cat_el = item.find("statusCategory")

        t = TicketSnapshot(
            ticket_id=ticket_id,
            key=key,
            key_id_raw=key_id_raw,
            file_name=path.name,
            export_date=export_date,
            link=get_child_raw(item, "link"),
            project_key=project_key,
            project_name=project_name,
            title=get_child_text(item, "title"),
            summary=get_child_text(item, "summary"),
            issue_type=get_child_text(item, "type"),
            priority=get_child_text(item, "priority"),
            status=get_child_text(item, "status"),
            status_category=attr(status_cat_el, "key"),
            resolution=get_child_text(item, "resolution"),
            assignee=get_child_text(item, "assignee"),
            reporter=get_child_text(item, "reporter"),
            created=parse_jira_date(text_of(item.find("created"))),
            updated=parse_jira_date(text_of(item.find("updated"))),
            due=parse_jira_date(text_of(item.find("due"))),
            labels=extract_labels(item),
            sprint=get_cf(customfields, "Sprint"),
            theme=get_cf(customfields, "Theme"),
            risk_level=get_cf(customfields, "Risk level"),
            risk_priority=get_cf(customfields, "Risk priority"),
            target_type=get_cf(customfields, "Target Type"),
            taille_demande=get_cf(customfields, "Taille de la demande"),
            horizon=get_cf(customfields, "Horizon"),
            sla_applicable=get_cf(customfields, "SLA applicable"),
            sync_billable_part=get_cf(customfields, "Sync billable part"),
            description_text=description_text,
            acceptance_criteria_text=ac_text,
            acceptance_criteria_source=ac_source,
            has_acceptance_criteria=has_ac,
            comments_count=comments_count,
            comments_signature=comments_signature,
            attachments_count=attachments_count,
            attachments_signature=attachments_signature,
            issue_links_count=links_count,
            issue_links_signature=links_signature,
            subtasks_count=extract_subtasks_count(item),
            customfields_signature=important_customfields_signature(customfields),
            customfields=customfields,
        )
        tickets[ticket_id] = t

    for ticket_id, count in seen_ids.items():
        if count > 1:
            warnings.append(f"Doublon ticket_id={ticket_id} dans {path.name}; derniere occurrence conservee.")

    return SnapshotFile(path=path, export_date=export_date, source_date_text=source_date_text, tickets=tickets, warnings=warnings)


# =============================================================================
# Comparaison des snapshots
# =============================================================================

def effective(after: Optional[TicketSnapshot], before: Optional[TicketSnapshot], attr_name: str) -> str:
    if after is not None:
        value = getattr(after, attr_name, "")
        if value:
            return str(value)
    if before is not None:
        value = getattr(before, attr_name, "")
        if value:
            return str(value)
    return ""


def base_event(
    prev: SnapshotFile,
    curr: SnapshotFile,
    before: Optional[TicketSnapshot],
    after: Optional[TicketSnapshot],
    field_name: str,
    change_type: str,
    before_value: str,
    after_value: str,
    before_length: Optional[int] = None,
    after_length: Optional[int] = None,
) -> ChangeEvent:
    ticket_id = after.ticket_id if after else before.ticket_id  # type: ignore[union-attr]
    return ChangeEvent(
        window_start=prev.export_date,
        window_end=curr.export_date,
        previous_file=prev.path.name,
        current_file=curr.path.name,
        ticket_id=ticket_id,
        key_before=before.key if before else "",
        key_after=after.key if after else "",
        project_key=effective(after, before, "project_key"),
        sprint_before=before.sprint if before else "",
        sprint_after=after.sprint if after else "",
        sprint_effective=effective(after, before, "sprint"),
        status_before=before.status if before else "",
        status_after=after.status if after else "",
        status_effective=effective(after, before, "status"),
        status_category_before=before.status_category if before else "",
        status_category_after=after.status_category if after else "",
        issue_type_before=before.issue_type if before else "",
        issue_type_after=after.issue_type if after else "",
        issue_type_effective=effective(after, before, "issue_type"),
        assignee_before=before.assignee if before else "",
        assignee_after=after.assignee if after else "",
        priority_before=before.priority if before else "",
        priority_after=after.priority if after else "",
        field_name=field_name,
        change_type=change_type,
        before_value=before_value,
        after_value=after_value,
        before_length=before_length,
        after_length=after_length,
        before_excerpt=excerpt(before_value),
        after_excerpt=excerpt(after_value),
        jira_updated_before=before.updated if before else None,
        jira_updated_after=after.updated if after else None,
        link=after.link if after and after.link else (before.link if before else ""),
    )


def compare_text_field(
    prev: SnapshotFile,
    curr: SnapshotFile,
    before: TicketSnapshot,
    after: TicketSnapshot,
    attr_name: str,
    field_label: str,
    change_label: str,
) -> Optional[ChangeEvent]:
    before_text = canonical_text(getattr(before, attr_name, ""))
    after_text = canonical_text(getattr(after, attr_name, ""))
    if stable_hash(before_text) == stable_hash(after_text):
        return None
    return base_event(
        prev,
        curr,
        before,
        after,
        field_label,
        change_label,
        before_text,
        after_text,
        len(before_text),
        len(after_text),
    )


def compare_snapshots(prev: SnapshotFile, curr: SnapshotFile) -> List[ChangeEvent]:
    events: List[ChangeEvent] = []
    before_ids = set(prev.tickets.keys())
    after_ids = set(curr.tickets.keys())

    for ticket_id in sorted(after_ids - before_ids):
        after = curr.tickets[ticket_id]
        events.append(base_event(
            prev, curr, None, after,
            "Ticket", "Ticket apparu dans l'export",
            "", after.summary or after.key,
            None, len(after.summary or after.key),
        ))

    for ticket_id in sorted(before_ids - after_ids):
        before = prev.tickets[ticket_id]
        events.append(base_event(
            prev, curr, before, None,
            "Ticket", "Ticket disparu de l'export",
            before.summary or before.key, "",
            len(before.summary or before.key), None,
        ))

    for ticket_id in sorted(before_ids & after_ids):
        before = prev.tickets[ticket_id]
        after = curr.tickets[ticket_id]

        desc_event = compare_text_field(prev, curr, before, after, "description_text", "Description", "Description modifiée")
        if desc_event:
            events.append(desc_event)

        ac_event = compare_text_field(prev, curr, before, after, "acceptance_criteria_text", "Acceptance Criteria", "Acceptance Criteria modifiés")
        if ac_event:
            events.append(ac_event)

        for attr_name, label in TRACKED_SIMPLE_FIELDS:
            before_value = canonical_text(str(getattr(before, attr_name, "") or ""))
            after_value = canonical_text(str(getattr(after, attr_name, "") or ""))
            if before_value != after_value:
                events.append(base_event(
                    prev, curr, before, after,
                    label, f"{label} modifié",
                    before_value, after_value,
                    len(before_value), len(after_value),
                ))

        # Signaux de tracabilite.
        if before.comments_signature != after.comments_signature:
            if after.comments_count > before.comments_count:
                change_type = "Commentaire ajouté"
            elif after.comments_count < before.comments_count:
                change_type = "Commentaire supprimé"
            else:
                change_type = "Commentaires modifiés"
            events.append(base_event(
                prev, curr, before, after,
                "Commentaires", change_type,
                str(before.comments_count), str(after.comments_count),
                before.comments_count, after.comments_count,
            ))

        if before.attachments_signature != after.attachments_signature:
            if after.attachments_count > before.attachments_count:
                change_type = "Pièce jointe ajoutée"
            elif after.attachments_count < before.attachments_count:
                change_type = "Pièce jointe supprimée"
            else:
                change_type = "Pièces jointes modifiées"
            events.append(base_event(
                prev, curr, before, after,
                "Pièces jointes", change_type,
                str(before.attachments_count), str(after.attachments_count),
                before.attachments_count, after.attachments_count,
            ))

        if before.issue_links_signature != after.issue_links_signature:
            events.append(base_event(
                prev, curr, before, after,
                "Liens Jira", "Liens Jira modifiés",
                str(before.issue_links_count), str(after.issue_links_count),
                before.issue_links_count, after.issue_links_count,
            ))

        if before.subtasks_count != after.subtasks_count:
            events.append(base_event(
                prev, curr, before, after,
                "Sous-tâches", "Sous-tâches modifiées",
                str(before.subtasks_count), str(after.subtasks_count),
                before.subtasks_count, after.subtasks_count,
            ))

        # Changement customfield global non deja capte explicitement.
        if before.customfields_signature != after.customfields_signature:
            # On liste les noms des customfields differents pour donner de la valeur.
            changed_names = []
            all_names = sorted(set(before.customfields) | set(after.customfields), key=str.lower)
            for name in all_names:
                if canonical_text(before.customfields.get(name, "")) != canonical_text(after.customfields.get(name, "")):
                    if name not in {"Acceptance Criteria", "Acceptance criteria", "Story Description", "Sprint", "Theme", "Risk level", "Risk priority", "Target Type", "Taille de la demande", "Horizon", "SLA applicable"}:
                        changed_names.append(name)
            if changed_names:
                events.append(base_event(
                    prev, curr, before, after,
                    "Custom fields", "Custom fields modifiés",
                    "; ".join(changed_names[:30]),
                    "; ".join(changed_names[:30]),
                    len(changed_names), len(changed_names),
                ))

    return events


def intervals_overlap(start_a: datetime, end_a: datetime, start_b: Optional[datetime], end_b: Optional[datetime]) -> bool:
    if start_b is None and end_b is None:
        return True
    if start_b is None:
        return start_a <= end_b  # type: ignore[operator]
    if end_b is None:
        return end_a >= start_b
    return start_a <= end_b and end_a >= start_b


def norm_filter(value: str) -> str:
    return normalize_for_check(value or "")


def contains_filter(value: str, pattern: Optional[str]) -> bool:
    if not pattern:
        return True
    return norm_filter(pattern) in norm_filter(value)


def event_matches_filters(
    event: ChangeEvent,
    sprint: Optional[str] = None,
    status: Optional[str] = None,
    issue_type: Optional[str] = None,
    project: Optional[str] = None,
    assignee: Optional[str] = None,
    field_name: Optional[str] = None,
) -> bool:
    return (
        contains_filter(event.sprint_effective, sprint)
        and contains_filter(event.status_effective, status)
        and contains_filter(event.issue_type_effective, issue_type)
        and contains_filter(event.project_key, project)
        and (contains_filter(event.assignee_after or event.assignee_before, assignee))
        and contains_filter(event.field_name, field_name)
    )


# =============================================================================
# KPI et lignes de reporting
# =============================================================================

def count_unique_tickets(events: Sequence[ChangeEvent]) -> int:
    return len({e.ticket_id for e in events})


def latest_snapshot_rows(snapshot: Optional[SnapshotFile], report_dt: datetime) -> List[Dict[str, Any]]:
    if not snapshot:
        return []
    rows = []
    for t in sorted(snapshot.tickets.values(), key=lambda x: (x.project_key, x.key)):
        rows.append({
            "Ticket ID": t.ticket_id,
            "Key": t.key,
            "Projet": t.project_key,
            "Sprint": t.sprint,
            "Statut": t.status,
            "Catégorie statut": t.status_category,
            "Type": t.issue_type,
            "Priorité": t.priority,
            "Assigné": t.assignee,
            "Résumé": t.summary,
            "Created": excel_dt(t.created),
            "Updated": excel_dt(t.updated),
            "Âge jours": days_between(t.created, report_dt),
            "Jours depuis update": days_between(t.updated, report_dt),
            "AC présent": t.has_acceptance_criteria,
            "Source AC": t.acceptance_criteria_source,
            "Commentaires": t.comments_count,
            "Pièces jointes": t.attachments_count,
            "Liens Jira": t.issue_links_count,
            "Lien": t.link,
        })
    return rows


def snapshot_rows(snapshots: Sequence[SnapshotFile]) -> List[Dict[str, Any]]:
    rows = []
    for idx, snap in enumerate(snapshots, start=1):
        rows.append({
            "Ordre": idx,
            "Date export": excel_dt(snap.export_date),
            "Date source": snap.source_date_text,
            "Fichier": snap.path.name,
            "Chemin": str(snap.path),
            "Tickets": len(snap.tickets),
            "Warnings": " | ".join(snap.warnings),
        })
    return rows


def kpi_by_window(events: Sequence[ChangeEvent]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[datetime, datetime, str, str], List[ChangeEvent]] = defaultdict(list)
    for e in events:
        grouped[(e.window_start, e.window_end, e.previous_file, e.current_file)].append(e)

    rows = []
    for (start, end, prev_file, curr_file), items in sorted(grouped.items(), key=lambda kv: kv[0][1]):
        rows.append({
            "Fenêtre début": excel_dt(start),
            "Fenêtre fin / détection": excel_dt(end),
            "Fichier précédent": prev_file,
            "Fichier courant": curr_file,
            "Tickets modifiés": count_unique_tickets(items),
            "Total modifications": len(items),
            "Descriptions modifiées": sum(e.field_name == "Description" for e in items),
            "AC modifiés": sum(e.field_name == "Acceptance Criteria" for e in items),
            "Changements statut": sum(e.field_name == "Statut" for e in items),
            "Changements sprint": sum(e.field_name == "Sprint" for e in items),
            "Commentaires ajoutés/modifiés": sum(e.field_name == "Commentaires" for e in items),
            "Tickets apparus": sum(e.change_type == "Ticket apparu dans l'export" for e in items),
            "Tickets disparus": sum(e.change_type == "Ticket disparu de l'export" for e in items),
        })
    return rows


def kpi_by_field(events: Sequence[ChangeEvent]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[ChangeEvent]] = defaultdict(list)
    for e in events:
        grouped[e.field_name].append(e)
    rows = []
    for field_name, items in sorted(grouped.items(), key=lambda kv: (-len(kv[1]), kv[0].lower())):
        rows.append({
            "Champ modifié": field_name,
            "Modifications": len(items),
            "Tickets concernés": count_unique_tickets(items),
            "Sprint le plus fréquent": most_common([e.sprint_effective for e in items]),
            "Statut le plus fréquent": most_common([e.status_effective for e in items]),
        })
    return rows


def kpi_by_sprint_status_type(events: Sequence[ChangeEvent]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[str, str, str, str], List[ChangeEvent]] = defaultdict(list)
    for e in events:
        grouped[(e.sprint_effective or "(vide)", e.status_effective or "(vide)", e.issue_type_effective or "(vide)", e.field_name)].append(e)
    rows = []
    for (sprint, status, issue_type, field_name), items in sorted(grouped.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        rows.append({
            "Sprint": sprint,
            "Statut": status,
            "Type": issue_type,
            "Champ modifié": field_name,
            "Modifications": len(items),
            "Tickets concernés": count_unique_tickets(items),
        })
    return rows


def top_modified_tickets(events: Sequence[ChangeEvent]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[ChangeEvent]] = defaultdict(list)
    for e in events:
        grouped[e.ticket_id].append(e)
    rows = []
    for ticket_id, items in sorted(grouped.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        latest = max(items, key=lambda e: e.window_end)
        rows.append({
            "Ticket ID": ticket_id,
            "Key": latest.key_after or latest.key_before,
            "Projet": latest.project_key,
            "Sprint effectif": latest.sprint_effective,
            "Statut effectif": latest.status_effective,
            "Type effectif": latest.issue_type_effective,
            "Assigné": latest.assignee_after or latest.assignee_before,
            "Modifications": len(items),
            "Champs modifiés": "; ".join(sorted(set(e.field_name for e in items))),
            "Première détection": excel_dt(min(e.window_end for e in items)),
            "Dernière détection": excel_dt(max(e.window_end for e in items)),
            "Lien": latest.link,
        })
    return rows


def lifecycle_events(events: Sequence[ChangeEvent]) -> List[Dict[str, Any]]:
    return [e.as_row() for e in events if e.field_name == "Ticket"]


def most_common(values: Iterable[str]) -> str:
    vals = [v for v in values if v]
    if not vals:
        return ""
    return Counter(vals).most_common(1)[0][0]


def limitations_rows(snapshots: Sequence[SnapshotFile], period_start: Optional[datetime], period_end: Optional[datetime]) -> List[Dict[str, Any]]:
    rows = [
        {
            "Sujet": "Nature de la date",
            "Message": "Les dates de modification sont des fenêtres de détection entre deux exports RSS. Le RSS ne donne pas l'heure exacte ni l'auteur exact de chaque changement de champ.",
        },
        {
            "Sujet": "Identifiant ticket",
            "Message": "La comparaison utilise l'id technique de <key id=...>. Si absent, le script utilise la key Jira et l'indique dans les warnings.",
        },
        {
            "Sujet": "Tickets disparus",
            "Message": "Un ticket disparu signifie qu'il est absent de l'export suivant. Cela peut venir d'une clôture, d'un changement de filtre Jira, d'un droit d'accès ou d'une suppression.",
        },
        {
            "Sujet": "Filtres Jira",
            "Message": "Pour des comparaisons fiables, tous les fichiers du répertoire doivent provenir du même filtre Jira ou d'un périmètre fonctionnel comparable.",
        },
    ]
    if period_start or period_end:
        rows.append({
            "Sujet": "Période demandée",
            "Message": f"Période appliquée sur l'intervalle de détection: {format_dt(period_start) if period_start else '-inf'} à {format_dt(period_end) if period_end else '+inf'}.",
        })
    for snap in snapshots:
        for w in snap.warnings:
            rows.append({"Sujet": f"Warning {snap.path.name}", "Message": w})
    return rows


# =============================================================================
# Excel
# =============================================================================

def import_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        return Workbook, BarChart, LineChart, Reference, Alignment, Border, Font, PatternFill, Side, get_column_letter
    except Exception as exc:
        raise SystemExit("openpyxl est requis. Lance: pip install openpyxl") from exc


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    return name[:31]


def write_sheet(ws: Any, rows: List[Dict[str, Any]], title_fill: str = "1F4E78") -> None:
    _, _, _, _, Alignment, Border, Font, PatternFill, Side, get_column_letter = import_openpyxl()
    ws.sheet_view.showGridLines = False
    if not rows:
        ws.append(["Aucune donnée"])
        ws["A1"].font = Font(bold=True)
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])

    header_fill = PatternFill("solid", fgColor=title_fill)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif isinstance(cell.value, float):
                cell.number_format = "0.00"

    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for cell in ws[col_letter][1: min(ws.max_row, 250)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, 11), 46)
        if any(x in header.lower() for x in ["extrait", "valeur", "résumé", "message", "lien", "chemin"]):
            width = min(max(width, 28), 70)
        ws.column_dimensions[col_letter].width = width


def set_cell(ws: Any, cell: str, value: Any, font: Any = None, fill: Any = None, alignment: Any = None, border: Any = None) -> None:
    ws[cell] = value
    if font:
        ws[cell].font = font
    if fill:
        ws[cell].fill = fill
    if alignment:
        ws[cell].alignment = alignment
    if border:
        ws[cell].border = border


def build_summary_sheet(
    wb: Any,
    sheet_name: str,
    all_events: Sequence[ChangeEvent],
    focus_events: Sequence[ChangeEvent],
    snapshots: Sequence[SnapshotFile],
    period_start: Optional[datetime],
    period_end: Optional[datetime],
    filters: Dict[str, Optional[str]],
) -> None:
    Workbook, BarChart, LineChart, Reference, Alignment, Border, Font, PatternFill, Side, get_column_letter = import_openpyxl()
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    dark = "17365D"
    blue = "1F4E78"
    light_blue = "D9EAF7"
    green = "E2F0D9"
    orange = "FCE4D6"
    grey = "F2F2F2"
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=18)
    subtitle_font = Font(color="1F4E78", bold=True, size=12)
    normal_bold = Font(bold=True)
    small_grey = Font(color="666666", italic=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:N2")
    set_cell(ws, "A1", "Suivi des modifications Jira RSS", title_font, PatternFill("solid", fgColor=dark), center)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8

    first_export = snapshots[0].export_date if snapshots else None
    last_export = snapshots[-1].export_date if snapshots else None
    focus_note = "Aucun filtre sprint/statut/type" if not any(filters.values()) else "; ".join([f"{k}={v}" for k, v in filters.items() if v])

    meta = [
        ("Exports analysés", len(snapshots)),
        ("Premier export", excel_dt(first_export)),
        ("Dernier export", excel_dt(last_export)),
        ("Période début", excel_dt(period_start)),
        ("Période fin", excel_dt(period_end)),
        ("Filtre focus", focus_note),
    ]
    start_row = 4
    for idx, (label, value) in enumerate(meta):
        r = start_row + idx
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = normal_bold
        ws[f"B{r}"] = value
        ws[f"A{r}"].fill = PatternFill("solid", fgColor=grey)
        ws[f"A{r}"].border = border
        ws[f"B{r}"].border = border
        if isinstance(value, datetime):
            ws[f"B{r}"].number_format = "yyyy-mm-dd hh:mm"

    # KPI cards.
    card_data = [
        ("Tickets modifiés", count_unique_tickets(focus_events), green),
        ("Total modifications", len(focus_events), light_blue),
        ("Descriptions modifiées", sum(e.field_name == "Description" for e in focus_events), orange),
        ("AC modifiés", sum(e.field_name == "Acceptance Criteria" for e in focus_events), orange),
        ("Changements statut", sum(e.field_name == "Statut" for e in focus_events), light_blue),
        ("Changements sprint", sum(e.field_name == "Sprint" for e in focus_events), light_blue),
        ("Tickets apparus", sum(e.change_type == "Ticket apparu dans l'export" for e in focus_events), green),
        ("Tickets disparus", sum(e.change_type == "Ticket disparu de l'export" for e in focus_events), orange),
    ]
    card_positions = ["D4", "G4", "J4", "M4", "D8", "G8", "J8", "M8"]
    for (label, value, fill_color), pos in zip(card_data, card_positions):
        col = ws[pos].column
        row = ws[pos].row
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        top = ws.cell(row=row, column=col)
        bottom = ws.cell(row=row + 1, column=col)
        top.value = label
        top.fill = PatternFill("solid", fgColor=blue)
        top.font = white_font
        top.alignment = center
        bottom.value = value
        bottom.fill = PatternFill("solid", fgColor=fill_color)
        bottom.font = Font(bold=True, size=18, color="17365D")
        bottom.alignment = center
        for rr in range(row, row + 3):
            for cc in range(col, col + 2):
                ws.cell(rr, cc).border = border

    # Petite phrase de lecture.
    ws.merge_cells("A12:N13")
    ws["A12"] = (
        "Lecture : chaque modification est une différence détectée entre deux exports. "
        "Les colonnes Sprint/Statut/Type permettent de filtrer ou trier rapidement les événements. "
        "Le focus ci-dessus applique les filtres passés en ligne de commande."
    )
    ws["A12"].font = small_grey
    ws["A12"].alignment = left

    # Table par champ.
    field_rows = kpi_by_field(focus_events)
    ws["A16"] = "Modifications par champ - focus"
    ws["A16"].font = subtitle_font
    field_headers = ["Champ modifié", "Modifications", "Tickets concernés"]
    for c, h in enumerate(field_headers, start=1):
        cell = ws.cell(17, c)
        cell.value = h
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = white_font
        cell.alignment = center
        cell.border = border
    for r_idx, row in enumerate(field_rows[:12], start=18):
        ws.cell(r_idx, 1).value = row["Champ modifié"]
        ws.cell(r_idx, 2).value = row["Modifications"]
        ws.cell(r_idx, 3).value = row["Tickets concernés"]
        for c in range(1, 4):
            ws.cell(r_idx, c).border = border

    if field_rows:
        chart = BarChart()
        chart.title = "Modifications par champ"
        chart.y_axis.title = "Nombre"
        chart.x_axis.title = "Champ"
        max_row = 17 + min(len(field_rows), 12)
        data = Reference(ws, min_col=2, min_row=17, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=18, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, "E16")

    # Evolution par fenetre.
    window_rows = kpi_by_window(focus_events)
    base_row = 35
    ws[f"A{base_row}"] = "Évolution par fenêtre d'export - focus"
    ws[f"A{base_row}"].font = subtitle_font
    headers = ["Fin fenêtre", "Total modifications", "Tickets modifiés", "Descriptions", "AC", "Statut"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(base_row + 1, c)
        cell.value = h
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = white_font
        cell.alignment = center
        cell.border = border
    for r_idx, row in enumerate(window_rows, start=base_row + 2):
        ws.cell(r_idx, 1).value = row["Fenêtre fin / détection"]
        ws.cell(r_idx, 2).value = row["Total modifications"]
        ws.cell(r_idx, 3).value = row["Tickets modifiés"]
        ws.cell(r_idx, 4).value = row["Descriptions modifiées"]
        ws.cell(r_idx, 5).value = row["AC modifiés"]
        ws.cell(r_idx, 6).value = row["Changements statut"]
        for c in range(1, 7):
            ws.cell(r_idx, c).border = border
        ws.cell(r_idx, 1).number_format = "yyyy-mm-dd hh:mm"

    if window_rows:
        chart2 = LineChart()
        chart2.title = "Évolution des modifications"
        chart2.y_axis.title = "Nombre"
        chart2.x_axis.title = "Date de détection"
        max_row = base_row + 1 + len(window_rows)
        data = Reference(ws, min_col=2, max_col=3, min_row=base_row + 1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=base_row + 2, max_row=max_row)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 8
        chart2.width = 16
        ws.add_chart(chart2, f"H{base_row}")

    # Mise en forme globale.
    for col_idx in range(1, 15):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "A16"


def build_excel_report(
    output_path: Path,
    snapshots: Sequence[SnapshotFile],
    all_period_events: Sequence[ChangeEvent],
    focus_events: Sequence[ChangeEvent],
    period_start: Optional[datetime],
    period_end: Optional[datetime],
    filters: Dict[str, Optional[str]],
) -> None:
    Workbook, *_ = import_openpyxl()
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_summary_sheet(wb, "Synthese", all_period_events, focus_events, snapshots, period_start, period_end, filters)

    # Onglets detail. On garde les deux niveaux : tout le perimetre periode, puis focus filtre.
    sheets = {
        "Modifications_focus": [e.as_row() for e in focus_events],
        "Modifications_toutes": [e.as_row() for e in all_period_events],
        "KPI_fenetres": kpi_by_window(focus_events),
        "KPI_sprint_status": kpi_by_sprint_status_type(focus_events),
        "Descriptions_modifiees": [e.as_row() for e in focus_events if e.field_name == "Description"],
        "AC_modifies": [e.as_row() for e in focus_events if e.field_name == "Acceptance Criteria"],
        "Lifecycle": lifecycle_events(focus_events),
        "Top_tickets": top_modified_tickets(focus_events),
        "Snapshots": snapshot_rows(snapshots),
        "Latest_snapshot": latest_snapshot_rows(snapshots[-1] if snapshots else None, snapshots[-1].export_date if snapshots else datetime.now(timezone.utc)),
        "Limites": limitations_rows(snapshots, period_start, period_end),
    }

    for name, rows in sheets.items():
        ws = wb.create_sheet(safe_sheet_name(name))
        write_sheet(ws, rows)

    wb.properties.creator = "analyse_jira_rss.py"
    wb.properties.title = "KPI modifications Jira RSS"
    wb.save(output_path)

    validate_xlsx_no_tables_or_autofilter(output_path)


def validate_xlsx_no_tables_or_autofilter(output_path: Path) -> None:
    """Evite les soucis Excel vus avec /xl/tables/table*.xml et autoFilter."""
    with zipfile.ZipFile(output_path, "r") as zf:
        names = zf.namelist()
        table_files = [n for n in names if n.startswith("xl/tables/")]
        if table_files:
            raise SystemExit(f"Validation KO: le fichier contient encore xl/tables/: {table_files[:5]}")
        offenders = []
        for n in names:
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                content = zf.read(n)
                if b"<autoFilter" in content or b"<tableParts" in content:
                    offenders.append(n)
        if offenders:
            raise SystemExit(f"Validation KO: autoFilter/tableParts trouves dans {offenders[:5]}")


# =============================================================================
# Orchestration
# =============================================================================

def load_snapshots(input_path: Path) -> List[SnapshotFile]:
    files = find_xml_files(input_path)
    snapshots = [parse_snapshot_file(path) for path in files]
    snapshots.sort(key=lambda s: (s.export_date, s.path.name))

    # Si deux fichiers ont exactement la meme date, on garde quand meme l'ordre par nom,
    # mais on le signale car la fenetre de modification sera ambigue.
    date_counts = Counter(s.export_date for s in snapshots)
    for snap in snapshots:
        if date_counts[snap.export_date] > 1:
            snap.warnings.append(f"Plusieurs exports ont la meme date exacte: {format_dt(snap.export_date)}. Ordre secondaire = nom de fichier.")
    return snapshots


def build_all_events(snapshots: Sequence[SnapshotFile]) -> List[ChangeEvent]:
    events: List[ChangeEvent] = []
    for prev, curr in zip(snapshots, snapshots[1:]):
        events.extend(compare_snapshots(prev, curr))
    return events


def filter_by_period(events: Sequence[ChangeEvent], period_start: Optional[datetime], period_end: Optional[datetime]) -> List[ChangeEvent]:
    return [e for e in events if intervals_overlap(e.window_start, e.window_end, period_start, period_end)]


def print_console_summary(snapshots: Sequence[SnapshotFile], all_period_events: Sequence[ChangeEvent], focus_events: Sequence[ChangeEvent], output_path: Path) -> None:
    print("\nRésumé rapide")
    print("--------------")
    print(f"Exports analysés          : {len(snapshots)}")
    if snapshots:
        print(f"Premier export            : {format_dt(snapshots[0].export_date)}")
        print(f"Dernier export            : {format_dt(snapshots[-1].export_date)}")
        print(f"Tickets dernier export    : {len(snapshots[-1].tickets)}")
    print(f"Modifications période     : {len(all_period_events)}")
    print(f"Modifications focus       : {len(focus_events)}")
    print(f"Tickets modifiés focus    : {count_unique_tickets(focus_events)}")
    print(f"Descriptions modifiées    : {sum(e.field_name == 'Description' for e in focus_events)}")
    print(f"AC modifiés               : {sum(e.field_name == 'Acceptance Criteria' for e in focus_events)}")
    print(f"Rapport                   : {output_path}")
    print("\nNote : les dates de modification sont des fenêtres entre deux exports RSS, pas un changelog exact Jira.")


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyse les modifications entre exports RSS/XML Jira et génère un Excel KPI.")
    parser.add_argument("input", help="Répertoire contenant les XML Jira, ou fichier XML unique.")
    parser.add_argument("--output", "-o", default="jira_kpi_report.xlsx", help="Fichier Excel de sortie.")
    parser.add_argument("--period-start", default=None, help="Début de période, ex: 2026-06-01 ou 2026-06-01 08:00:00+00:00.")
    parser.add_argument("--period-end", default=None, help="Fin de période, ex: 2026-06-05. Inclusif si date seule.")
    parser.add_argument("--sprint", default=None, help="Filtre focus sprint. Match partiel insensible à la casse, ex: Sprint 16.")
    parser.add_argument("--status", default=None, help="Filtre focus statut. Match partiel, ex: Ouvert.")
    parser.add_argument("--type", dest="issue_type", default=None, help="Filtre focus type Jira, ex: Bug, Story.")
    parser.add_argument("--project", default=None, help="Filtre focus projet, ex: PPMG.")
    parser.add_argument("--assignee", default=None, help="Filtre focus assigné.")
    parser.add_argument("--field", default=None, help="Filtre focus champ modifié, ex: Description, Acceptance Criteria, Statut.")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv)
    input_path = Path(args.input)
    output_path = Path(args.output)

    period_start = parse_period_boundary(args.period_start, end_of_day=False)
    period_end = parse_period_boundary(args.period_end, end_of_day=True)

    snapshots = load_snapshots(input_path)
    if len(snapshots) < 2:
        print("Attention : un seul export trouvé. Le rapport sera généré, mais aucune modification inter-snapshot ne peut être détectée.")

    all_events = build_all_events(snapshots)
    all_period_events = filter_by_period(all_events, period_start, period_end)

    filters = {
        "Sprint": args.sprint,
        "Statut": args.status,
        "Type": args.issue_type,
        "Projet": args.project,
        "Assigné": args.assignee,
        "Champ": args.field,
    }
    focus_events = [
        e for e in all_period_events
        if event_matches_filters(
            e,
            sprint=args.sprint,
            status=args.status,
            issue_type=args.issue_type,
            project=args.project,
            assignee=args.assignee,
            field_name=args.field,
        )
    ]

    build_excel_report(output_path, snapshots, all_period_events, focus_events, period_start, period_end, filters)
    print_console_summary(snapshots, all_period_events, focus_events, output_path)


if __name__ == "__main__":
    main()
